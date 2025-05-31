from faker import Faker
faker = Faker("zh_CN")


import openpyxl

def get_name():
    '''
    :return: 随机生成一个中文名称
    '''
    return faker.name()


def get_phone():
    '''
    :return: 返回随机的手机号
    '''
    return faker.phone_number()

def get_email():
    '''

    :return: 随机获取电子邮箱
    '''
    return faker.email()


class ReadExcel():


    def read_xlsx_excel(self,excel_path, sheet_name):
        """
        :param url: 文件路径
        :param sheet_name: 文件中的sheet名称
        :return:  返回文件中的数据
        """
        # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
        workbook = openpyxl.load_workbook(excel_path)
        # 根据指定表名获取表格并得到对应的sheet对象
        sheet = workbook[sheet_name]
        print(sheet)
        # 定义列表存储表格数据
        data = []
        # 遍历表格的每一行
        for row in sheet.rows:
            # 定义表格存储每一行数据
            da = []
            # 从每一行中遍历每一个单元格
            for cell in row:
                # 将行数据存储到da列表
                da.append(cell.value)
            # 存储每一行数据
            data.append(da)
        # 返回数据
        return data




    def write_xlsx_excel_add(self,excel_path, sheet_name, two_dimensional_data):
        '''
        追加写入xlsx格式文件
        参数：
            excel_path:文件路径
            sheet_name:表名
            two_dimensional_data：将要写入表格的数据（二维列表）
        '''
        # 使用openpyxl加载指定路径的Excel文件并得到对应的workbook对象
        workbook = openpyxl.load_workbook(excel_path)
        # 根据指定表名获取表格并得到对应的sheet对象
        sheet = workbook[sheet_name]
        for tdd in two_dimensional_data:
            sheet.append(tdd)
        # 保存到指定位置
        workbook.save(excel_path)
        print("数据追加写入成功。追加的工作目录是: {}；追加的数据是: {}".format(sheet_name,two_dimensional_data))


    def write_xlsx_excel(self,excel_path, sheet_name, two_dimensional_data):
        '''
        写入xlsx格式文件
        参数：
            excel_path:文件路径
            sheet_name:表名
            two_dimensional_data：将要写入表格的数据（二维列表）
        '''
        # 创建工作簿对象
        workbook = openpyxl.Workbook()
        # 创建工作表对象
        sheet = workbook.active
        # 设置该工作表的名字
        sheet.title = sheet_name
        # 遍历表格的每一行
        for i in range(0, len(two_dimensional_data)):
            # 遍历表格的每一列
            for j in range(0, len(two_dimensional_data[i])):
                # 写入数据（注意openpyxl的行和列是从1开始的，和我们平时的认知是一样的）
                sheet.cell(row=i + 1, column=j + 1, value=str(two_dimensional_data[i][j]))
        # 保存到指定位置
        workbook.save(excel_path)
        print("数据写入成功。写入的工作目录是: {};写入的数据是: {}".format(sheet_name,two_dimensional_data))

    def add_sheets(self,sheets):
        """
        :param sheets: 多个工作目录的名称，以列表/元组形式传入
        :return:
        """

        # 创建一个新的 Excel 工作簿
        wb = openpyxl.Workbook()

        # 删除默认创建的 "Sheet"（可选）
        default_sheet = wb.active
        wb.remove(default_sheet)
        # 新增多个工作表
        # sheets = ["销售数据", "用户统计", "财务报告"]
        for sheet_name in sheets:
            wb.create_sheet(title=sheet_name)  # 新增工作表
        # 保存文件
        wb.save("../data/new_workbook.xlsx")




datas = [["万股",22,908,45112]]
biaotu = [["姓名","电话号码","电子邮件"]]
shuju = [
    [get_name(),get_phone(),get_email()]
]
read_excel = ReadExcel()
data = read_excel.read_xlsx_excel("../data/new_workbook.xlsx","用户统计")
# read_excel.write_xlsx_excel("../data/ceshisjiku.xlsx","李云飞测试数据1",biaotu)
# for i in range(100):
#     shuju = [
#         [get_name(), get_phone(), get_email()]
#     ]
#     read_excel.write_xlsx_excel_add(r"../data/new_workbook.xlsx","用户统计",shuju)

# read_excel.add_sheets(("销售数据", "用户统计", "财务报告"))
print(data)
#
# for i in data:
#     print(i)